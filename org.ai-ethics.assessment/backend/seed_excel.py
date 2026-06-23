import os
import sys
import openpyxl
from sqlalchemy.orm import Session

# Add current directory to path to import app modules correctly
sys.path.append(os.path.abspath(os.path.dirname(__file__)))

from app.database import SessionLocal, engine
from app import models

def seed_excel_data():
    # 1. Create tables if they do not exist (dropping diagnostic_types first to update schema)
    print("Creating tables if they do not exist...")
    try:
        models.DiagnosticType.__table__.drop(bind=engine, checkfirst=True)
        print("Dropped existing diagnostic_types table to update schema.")
    except Exception as e:
        print(f"Could not drop diagnostic_types table: {e}")
        
    models.Base.metadata.create_all(bind=engine)
    
    db = SessionLocal()
    try:
        # Clear existing data from these tables to avoid duplication and allow re-runs
        print("Clearing existing diagnostic types and factor descriptions...")
        db.query(models.DiagnosticType).delete()
        db.query(models.FactorDescription).delete()
        db.commit()

        # Load Workbook
        excel_path = os.path.join(os.path.dirname(__file__), "..", "..", "유형_(26.06.01)_송애리 검토.xlsx")
        print(f"Loading Excel file from: {os.path.abspath(excel_path)}")
        wb = openpyxl.load_workbook(excel_path)
        
        # -------------------------------------------------------------
        # Parse Sheet1: Diagnostic Types
        # -------------------------------------------------------------
        print("Parsing Sheet1 (Diagnostic Types)...")
        sheet1 = wb['Sheet1']
        rank_map = {"상": "H", "중": "M", "하": "L"}
        type_count = 0
        
        for r in range(2, sheet1.max_row + 1):
            num = sheet1.cell(r, 1).value
            if num is None:
                continue
            
            type_name = sheet1.cell(r, 2).value
            comb_str = sheet1.cell(r, 3).value
            desc = sheet1.cell(r, 4).value
            guideline = sheet1.cell(r, 5).value
            discussion_prompt = sheet1.cell(r, 6).value
            
            # Map combination string (e.g. '중 - 하 - 하') to code (e.g., 'MLL')
            parts = [p.strip() for p in comb_str.split("-")]
            code = "".join([rank_map[p] for p in parts])
            
            diag_type = models.DiagnosticType(
                code=code,
                name=type_name,
                description=desc,
                guideline=guideline,
                discussion_prompt=discussion_prompt
            )
            db.add(diag_type)
            type_count += 1
            
        print(f"Added {type_count} diagnostic types.")

        # -------------------------------------------------------------
        # Parse Sheet2: Factor Descriptions
        # -------------------------------------------------------------
        print("Parsing Sheet2 (Factor Descriptions)...")
        sheet2 = wb['Sheet2']
        factor_count = 0
        
        # In Sheet2:
        # Col 1: Level ('상', '중', '하')
        # Col 2: '위험 인식'
        # Col 3: '편익 인식'
        # Col 4: '사회정의'
        # Row 3 is '상', Row 4 is '중', Row 5 is '하'
        
        factors_mapping = {
            2: "risk",      # Col 2 -> 위험 인식
            3: "benefit",   # Col 3 -> 편익 인식
            4: "justice"    # Col 4 -> 사회정의
        }
        
        for r in [3, 4, 5]:
            rank_val = sheet2.cell(r, 1).value
            if not rank_val:
                continue
            
            for col_idx, factor_name in factors_mapping.items():
                desc_val = sheet2.cell(r, col_idx).value
                if desc_val:
                    # Clean trailing or leading whitespaces
                    desc_val = desc_val.strip()
                    
                    fd = models.FactorDescription(
                        factor=factor_name,
                        rank=rank_val,
                        description=desc_val
                    )
                    db.add(fd)
                    factor_count += 1
                    
        print(f"Added {factor_count} factor descriptions.")
        
        db.commit()
        print("Database seeding from Excel completed successfully!")
        
    except Exception as e:
        print(f"Error during excel seeding: {e}")
        db.rollback()
    finally:
        db.close()

if __name__ == "__main__":
    seed_excel_data()
